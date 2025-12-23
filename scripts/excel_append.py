import pandas as pd
from openpyxl import load_workbook
from datetime import date

EXCEL_PATH = "data/nse_bse_business_growth.xlsx"

def append_df(sheet_name: str, df: pd.DataFrame):
    df = df.copy()
    df["trade_date"] = date.today().strftime("%Y-%m-%d")

    try:
        book = load_workbook(EXCEL_PATH)

        if sheet_name in book.sheetnames:
            existing = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name)

            # prevent duplicate day append
            if "trade_date" in existing.columns:
                if df["trade_date"].iloc[0] in existing["trade_date"].astype(str).values:
                    print(f"⚠️ {sheet_name}: Data already exists for today")
                    return

            combined = pd.concat([existing, df], ignore_index=True)

            with pd.ExcelWriter(
                EXCEL_PATH,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="replace"
            ) as writer:
                writer.book = book
                combined.to_excel(writer, sheet_name=sheet_name, index=False)

        else:
            with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="a") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    except FileNotFoundError:
        with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
