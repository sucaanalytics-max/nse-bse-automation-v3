import pandas as pd
from openpyxl import load_workbook

EXCEL_PATH = "data/nse_bse_business_growth.xlsx"

def append_to_sheet(sheet_name: str, new_df: pd.DataFrame):
    """
    Appends new_df to an existing Excel sheet.
    If sheet does not exist, it will be created.
    """

    try:
        book = load_workbook(EXCEL_PATH)
        if sheet_name in book.sheetnames:
            existing_df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name)
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        else:
            combined_df = new_df

        with pd.ExcelWriter(
            EXCEL_PATH,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace"
        ) as writer:
            writer.book = book
            combined_df.to_excel(writer, sheet_name=sheet_name, index=False)

    except FileNotFoundError:
        # First run / file missing (safety)
        with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl") as writer:
            new_df.to_excel(writer, sheet_name=sheet_name, index=False)
